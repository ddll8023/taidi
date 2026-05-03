"""任务三结果导出服务。"""

import json
import os
import re
from datetime import datetime

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.schemas.common import ErrorCode
from app.schemas.task3 import (
    Reference,
    Task3ExportContentResponse,
    Task3LatestExportResponse,
    Task3SingleExportResponse,
    Task3WorkspaceExportResponse,
)
from app.services.task3.helpers import _parse_question_rounds
from app.services.task3.planner import process_task3_question
from app.services.task3.verifier import verify_answer_quality
from app.utils.exception import ServiceException
from app.utils.logger_config import setup_logger

logger = setup_logger(__name__)
TASK3_QUESTION_CODE_PATTERN = re.compile(r"^B2\d{3}$")


# ========== 公共入口函数 ==========


def export_result_3(questions: list[dict], db: Session):
    """按题目列表导出 result_3 结果文件。"""
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="结果汇总")
    ws.append(["编号", "问题", "SQL 查询语法", "回答"])

    all_results = []
    success_count = 0
    fail_count = 0

    for idx, item in enumerate(questions):
        question_id = item.get("id", f"C{idx + 1:03d}")
        question_text = item.get("question", "")

        rounds = _parse_question_rounds(question_text)
        merged_question_text = " ".join(item.get("Q", "") for item in rounds)

        logger.info("处理问题 %d/%d: %s", idx + 1, len(questions), question_id)

        try:
            _validate_task3_question_id(question_id)
            response = process_task3_question(
                question=merged_question_text,
                db=db,
                context={"question_id": question_id},
            )

            answer_content = response.answer

            references_list = []
            for ref in answer_content.references:
                ref_dict = {
                    "paper_path": ref.paper_path,
                    "text": ref.text[:500] if ref.text else "",
                }
                if ref.paper_image:
                    ref_dict["paper_image"] = ref.paper_image
                references_list.append(ref_dict)

            answer_data = {
                "content": answer_content.content,
            }
            if references_list:
                answer_data["references"] = references_list

            quality_check = verify_answer_quality(
                answer_content.content,
                answer_content.references,
                merged_question_text,
            ).model_dump(mode="json")

            qa_pairs = [{
                "Q": merged_question_text,
                "A": answer_data,
            }]
            qa_pairs = _remove_task3_answer_images(qa_pairs)

            result_item = {
                "id": question_id,
                "question": json.dumps(rounds, ensure_ascii=False),
                "sql": response.sql or "",
                "answer": qa_pairs,
                "quality": quality_check,
            }
            all_results.append(result_item)
            success_count += 1

            ws.append([
                question_id,
                json.dumps(rounds, ensure_ascii=False),
                response.sql or "",
                json.dumps(qa_pairs, ensure_ascii=False),
            ])

        except Exception as exc:
            logger.error("处理问题失败: question_id=%s, error=%s", question_id, str(exc))
            fail_count += 1

            result_item = {
                "id": question_id,
                "question": json.dumps(rounds, ensure_ascii=False),
                "sql": "",
                "answer": [{
                    "Q": merged_question_text,
                    "A": {"content": f"处理失败: {str(exc)}"},
                }],
                "error": str(exc),
            }
            all_results.append(result_item)

            ws.append([
                question_id,
                json.dumps(rounds, ensure_ascii=False),
                "",
                json.dumps(result_item["answer"], ensure_ascii=False),
            ])

    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    result_dir = os.path.join(os.getcwd(), "result")
    os.makedirs(result_dir, exist_ok=True)

    result_path = os.path.join(result_dir, "result_3.xlsx")
    wb.save(result_path)
    logger.info(
        "result_3.xlsx 已生成: %s, 成功=%d, 失败=%d",
        result_path,
        success_count,
        fail_count,
    )

    json_path = os.path.join(result_dir, "result_3.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    logger.info("result_3.json 已生成: %s", json_path)

    summary_path = os.path.join(result_dir, "result_3_summary.json")
    summary = {
        "total_questions": len(questions),
        "success_count": success_count,
        "fail_count": fail_count,
        "generated_at": datetime.now().isoformat(),
        "files": {
            "xlsx": result_path,
            "json": json_path,
        },
    }
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    return result_path


def export_single_question_result(
    question_id: str,
    question: str,
    db: Session,
):
    """导出单个问题的任务三结果。"""
    try:
        _validate_task3_question_id(question_id)
        response = process_task3_question(
            question=question,
            db=db,
            context={"question_id": question_id},
        )

        answer_content = response.answer

        references_list = []
        for ref in answer_content.references:
            ref_dict = {
                "paper_path": ref.paper_path,
                "text": ref.text[:500] if ref.text else "",
                "page_no": ref.page_no,
            }
            if ref.paper_image:
                ref_dict["paper_image"] = ref.paper_image
            references_list.append(ref_dict)

        return Task3SingleExportResponse(
            id=question_id,
            question=question,
            sql=response.sql,
            answer=Task3ExportContentResponse(
                content=answer_content.content,
                references=references_list,
            ),
            success=True,
            error=None,
        )

    except Exception as exc:
        logger.error("单问题导出失败: question_id=%s, error=%s", question_id, str(exc))
        return Task3SingleExportResponse(
            id=question_id,
            question=question,
            sql=None,
            answer=Task3ExportContentResponse(
                content=f"处理失败: {str(exc)}",
                references=[],
            ),
            success=False,
            error=str(exc),
        )


def format_reference_for_output(ref: Reference):
    """将引用对象格式化为导出结构。"""
    return {
        "paper_path": ref.paper_path,
        "text": ref.text,
        "page_no": ref.page_no,
        "paper_image": ref.paper_image,
    }


def validate_export_result(result: dict):
    """校验导出结果是否满足基础约束。"""
    validation = {
        "is_valid": True,
        "errors": [],
        "warnings": [],
    }

    if not result.get("id"):
        validation["errors"].append("缺少问题编号")
        validation["is_valid"] = False

    if not result.get("question"):
        validation["errors"].append("缺少问题内容")
        validation["is_valid"] = False

    answer = result.get("answer", [])
    if isinstance(answer, list):
        has_content = any(
            isinstance(item, dict)
            and isinstance(item.get("A"), dict)
            and bool(item["A"].get("content"))
            for item in answer
        )
    else:
        has_content = bool(answer.get("content")) if isinstance(answer, dict) else False

    if not has_content:
        validation["warnings"].append("答案内容为空")

    question = result.get("question", "")

    cause_keywords = ["原因", "是因为", "由于", "为何", "为什么"]
    has_cause_question = any(kw in question for kw in cause_keywords)

    if has_cause_question:
        if isinstance(answer, list):
            references = []
            for item in answer:
                if isinstance(item, dict) and isinstance(item.get("A"), dict):
                    references.extend(item["A"].get("references", []))
        else:
            references = answer.get("references", []) if isinstance(answer, dict) else []
        if not references:
            validation["warnings"].append("原因类问题缺少引用来源")

    return validation


def export_result_3_from_workspace(db: Session):
    """从当前工作台记录导出 result_3.xlsx。"""
    from app.models.task3_workspace import Task3Workspace
    from app.models.task3_question_item import Task3QuestionItem
    from sqlalchemy import select, func

    stmt = select(Task3Workspace).order_by(Task3Workspace.id.desc()).limit(1)
    workspace = db.execute(stmt).scalar_one_or_none()

    if workspace is None:
        raise ServiceException(ErrorCode.DATA_NOT_FOUND, "工作台不存在，请先导入附件6")

    stmt = select(Task3QuestionItem).where(
        Task3QuestionItem.workspace_id == workspace.id
    ).order_by(Task3QuestionItem.question_code)
    questions = list(db.execute(stmt).scalars().all())

    if not questions:
        raise ServiceException(ErrorCode.DATA_NOT_FOUND, "没有可导出的题目")

    workspace.total_questions = len(questions)
    workspace.answered_count = sum(1 for item in questions if item.status == 2)
    workspace.failed_count = sum(1 for item in questions if item.status == 3)
    workspace.pending_count = sum(1 for item in questions if item.status == 0)

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="结果汇总")
    ws.append(["编号", "问题", "SQL 查询语法", "回答"])

    success_count = 0
    fail_count = 0

    for question in questions:
        question_id = question.question_code
        question_text = question.question_raw_json or ""
        rounds = _parse_question_rounds(question_text)
        try:
            answer_json = _ensure_non_empty_qa_pairs(question_text, question.answer_json or [])
            answer_json = _remove_task3_answer_images(answer_json)

            ws.append([
                question_id,
                json.dumps(rounds, ensure_ascii=False),
                question.sql_text or "",
                json.dumps(answer_json, ensure_ascii=False),
            ])

            if question.status == 2:
                success_count += 1
            else:
                fail_count += 1

        except Exception as exc:
            logger.error("导出题目失败: question_id=%s, error=%s", question_id, str(exc))
            fail_count += 1

            ws.append([
                question_id,
                json.dumps(rounds, ensure_ascii=False),
                "",
                json.dumps(
                    [{
                        "Q": rounds[0].get("Q", "") if rounds else question_text,
                        "A": {"content": f"处理失败: {str(exc)}"},
                    }],
                    ensure_ascii=False,
                ),
            ])

    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    result_dir = os.path.join(os.getcwd(), "result")
    os.makedirs(result_dir, exist_ok=True)

    result_path = os.path.join(result_dir, "result_3.xlsx")
    wb.save(result_path)

    # 更新工作台导出信息
    workspace.last_export_path = result_path
    workspace.last_exported_at = datetime.now()
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.error("保存导出结果失败: %s", str(exc), exc_info=True)
        raise ServiceException(ErrorCode.INTERNAL_ERROR, "导出失败") from exc

    logger.info(
        "result_3.xlsx 导出完成: path=%s, success=%d, failed=%d",
        result_path,
        success_count,
        fail_count,
    )

    return Task3WorkspaceExportResponse(
        xlsx_path=result_path,
        success_count=success_count,
        fail_count=fail_count,
        total=len(questions),
    )


def get_latest_export_info(db: Session):
    """获取最近一次导出记录信息，无记录时返回带 message 的默认结构。"""
    from app.models.task3_workspace import Task3Workspace
    from sqlalchemy import select

    stmt = select(Task3Workspace).order_by(Task3Workspace.id.desc()).limit(1)
    workspace = db.execute(stmt).scalar_one_or_none()

    if workspace is None or not workspace.last_export_path:
        return {"message": "暂无导出记录"}

    return Task3LatestExportResponse(
        xlsx_path=workspace.last_export_path,
        exported_at=workspace.last_exported_at.isoformat() if workspace.last_exported_at else None,
        total_questions=workspace.total_questions,
        answered_count=workspace.answered_count,
    )


"""辅助函数"""


def _ensure_non_empty_qa_pairs(question_value, qa_pairs: list[dict]):
    """确保导出时至少存在一组问答结果。"""
    if qa_pairs:
        return qa_pairs

    rounds = _parse_question_rounds(question_value)
    first_question = rounds[0].get("Q", "") if rounds else str(question_value or "")
    return [
        {
            "Q": first_question,
            "A": {
                "content": "回答生成失败：未生成任何有效轮次结果，请重新执行该题。"
            },
        }
    ]


def _remove_task3_answer_images(qa_pairs: list[dict]):
    """移除任务三答案中的 image 字段。"""
    normalized_pairs = []
    for item in qa_pairs:
        if not isinstance(item, dict):
            normalized_pairs.append(item)
            continue

        normalized_item = dict(item)
        answer = normalized_item.get("A")
        if isinstance(answer, dict):
            normalized_answer = dict(answer)
            normalized_answer.pop("image", None)
            normalized_item["A"] = normalized_answer
        normalized_pairs.append(normalized_item)
    return normalized_pairs


def _validate_task3_question_id(question_id: str):
    """校验任务三题目编号格式。"""
    if not TASK3_QUESTION_CODE_PATTERN.fullmatch(str(question_id or "").strip()):
        raise ServiceException(ErrorCode.PARAM_ERROR, "任务三题目编号非法")

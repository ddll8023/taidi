import json
import os
from datetime import datetime

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.schemas.task3 import Reference
from app.services.task3_planner import process_task3_question
from app.services.task3_verifier import verify_answer_quality
from app.utils.logger_config import setup_logger

logger = setup_logger(__name__)

CHART_TYPE_MAP = {
    "line": "折线图",
    "bar": "柱状图",
    "pie": "饼图",
    "horizontal_bar": "条形图",
    "grouped_bar": "分组柱状图",
    "radar": "雷达图",
    "histogram": "直方图",
    "scatter": "散点图",
    "box": "箱线图",
}


def _parse_question_rounds(question_value) -> list[dict]:
    if isinstance(question_value, list):
        parsed = question_value
    else:
        try:
            parsed = json.loads(question_value)
        except (json.JSONDecodeError, TypeError):
            parsed = [{"Q": str(question_value or "")}]

    if not isinstance(parsed, list):
        parsed = [{"Q": str(parsed)}]

    rounds = []
    for item in parsed:
        if isinstance(item, dict):
            q_text = str(item.get("Q", "")).strip()
        else:
            q_text = str(item).strip()
        if q_text:
            rounds.append({"Q": q_text})

    if not rounds:
        rounds.append({"Q": str(question_value or "").strip()})
    return rounds


def _format_chart_type(chart_type: str | None) -> str:
    if not chart_type:
        return "无"
    return CHART_TYPE_MAP.get(chart_type, chart_type)


def _ensure_non_empty_qa_pairs(question_value, qa_pairs: list[dict]) -> list[dict]:
    if qa_pairs:
        return qa_pairs

    rounds = _parse_question_rounds(question_value)
    first_question = rounds[0].get("Q", "") if rounds else str(question_value or "")
    return [
        {
            "Q": first_question,
            "A": {
                "content": "回答生成失败：未生成任何有效轮次结果，请重新执行该题。"
            },
        }
    ]


def _convert_path_to_url(file_path: str) -> str:
    if not file_path:
        return ""
    filename = os.path.basename(file_path)
    return f"./result/{filename}"


def export_result_3(questions: list[dict], db: Session) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="结果汇总")
    ws.append(["编号", "问题", "SQL查询语句", "图形格式", "回答"])

    all_results = []
    success_count = 0
    fail_count = 0

    for idx, item in enumerate(questions):
        question_id = item.get("id", f"C{idx + 1:03d}")
        question_text = item.get("question", "")

        rounds = _parse_question_rounds(question_text)
        merged_question_text = " ".join(item.get("Q", "") for item in rounds)

        logger.info("处理问题 %d/%d: %s", idx + 1, len(questions), question_id)

        try:
            response = process_task3_question(
                question=merged_question_text,
                db=db,
                context={"question_id": question_id},
            )

            answer_content = response.answer

            image_paths = []
            if answer_content.image:
                for img in answer_content.image:
                    image_paths.append(_convert_path_to_url(img))

            references_list = []
            for ref in answer_content.references:
                ref_dict = {
                    "paper_path": ref.paper_path,
                    "text": ref.text[:500] if ref.text else "",
                }
                if ref.paper_image:
                    ref_dict["paper_image"] = ref.paper_image
                references_list.append(ref_dict)

            answer_data = {
                "content": answer_content.content,
            }
            if image_paths:
                answer_data["image"] = image_paths
            if references_list:
                answer_data["references"] = references_list

            quality_check = verify_answer_quality(
                answer_content.content,
                answer_content.references,
                merged_question_text,
            )

            qa_pairs = [{
                "Q": merged_question_text,
                "A": answer_data,
            }]
            chart_type = _format_chart_type(response.chart_type)

            result_item = {
                "id": question_id,
                "question": json.dumps(rounds, ensure_ascii=False),
                "sql": response.sql or "",
                "chart_type": chart_type,
                "answer": qa_pairs,
                "quality": quality_check,
            }
            all_results.append(result_item)
            success_count += 1

            ws.append([
                question_id,
                json.dumps(rounds, ensure_ascii=False),
                response.sql or "",
                chart_type,
                json.dumps(qa_pairs, ensure_ascii=False),
            ])

        except Exception as exc:
            logger.error("处理问题失败: question_id=%s, error=%s", question_id, str(exc))
            fail_count += 1

            result_item = {
                "id": question_id,
                "question": json.dumps(rounds, ensure_ascii=False),
                "sql": "",
                "chart_type": "无",
                "answer": [{
                    "Q": merged_question_text,
                    "A": {"content": f"处理失败: {str(exc)}"},
                }],
                "error": str(exc),
            }
            all_results.append(result_item)

            ws.append([
                question_id,
                json.dumps(rounds, ensure_ascii=False),
                "",
                "无",
                json.dumps(result_item["answer"], ensure_ascii=False),
            ])

    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    result_dir = os.path.join(os.getcwd(), "result")
    os.makedirs(result_dir, exist_ok=True)

    result_path = os.path.join(result_dir, "result_3.xlsx")
    wb.save(result_path)
    logger.info(
        "result_3.xlsx 已生成: %s, 成功=%d, 失败=%d",
        result_path,
        success_count,
        fail_count,
    )

    json_path = os.path.join(result_dir, "result_3.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    logger.info("result_3.json 已生成: %s", json_path)

    summary_path = os.path.join(result_dir, "result_3_summary.json")
    summary = {
        "total_questions": len(questions),
        "success_count": success_count,
        "fail_count": fail_count,
        "generated_at": datetime.now().isoformat(),
        "files": {
            "xlsx": result_path,
            "json": json_path,
        },
    }
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    return result_path


def export_single_question_result(
    question_id: str,
    question: str,
    db: Session,
) -> dict:
    try:
        response = process_task3_question(
            question=question,
            db=db,
            context={"question_id": question_id},
        )

        answer_content = response.answer

        image_paths = []
        if answer_content.image:
            for img in answer_content.image:
                image_paths.append(_convert_path_to_url(img))

        references_list = []
        for ref in answer_content.references:
            ref_dict = {
                "paper_path": ref.paper_path,
                "text": ref.text[:500] if ref.text else "",
                "page_no": ref.page_no,
            }
            if ref.paper_image:
                ref_dict["paper_image"] = ref.paper_image
            references_list.append(ref_dict)

        return {
            "id": question_id,
            "question": question,
            "sql": response.sql,
            "answer": {
                "content": answer_content.content,
                "image": image_paths,
                "references": references_list,
            },
            "success": True,
        }

    except Exception as exc:
        logger.error("单问题导出失败: question_id=%s, error=%s", question_id, str(exc))
        return {
            "id": question_id,
            "question": question,
            "sql": None,
            "answer": {"content": f"处理失败: {str(exc)}"},
            "success": False,
            "error": str(exc),
        }


def format_reference_for_output(ref: Reference) -> dict:
    return {
        "paper_path": ref.paper_path,
        "text": ref.text,
        "page_no": ref.page_no,
        "paper_image": ref.paper_image,
    }


def validate_export_result(result: dict) -> dict:
    validation = {
        "is_valid": True,
        "errors": [],
        "warnings": [],
    }

    if not result.get("id"):
        validation["errors"].append("缺少问题编号")
        validation["is_valid"] = False

    if not result.get("question"):
        validation["errors"].append("缺少问题内容")
        validation["is_valid"] = False

    answer = result.get("answer", [])
    if isinstance(answer, list):
        has_content = any(
            isinstance(item, dict)
            and isinstance(item.get("A"), dict)
            and bool(item["A"].get("content"))
            for item in answer
        )
    else:
        has_content = bool(answer.get("content")) if isinstance(answer, dict) else False

    if not has_content:
        validation["warnings"].append("答案内容为空")

    question = result.get("question", "")

    cause_keywords = ["原因", "是因为", "由于", "为何", "为什么"]
    has_cause_question = any(kw in question for kw in cause_keywords)

    if has_cause_question:
        if isinstance(answer, list):
            references = []
            for item in answer:
                if isinstance(item, dict) and isinstance(item.get("A"), dict):
                    references.extend(item["A"].get("references", []))
        else:
            references = answer.get("references", []) if isinstance(answer, dict) else []
        if not references:
            validation["warnings"].append("原因类问题缺少引用来源")

    return validation


# ─────────────────────────────────────────────────────────────────────────────
# 工作台导出（基于数据库题目记录）
# ─────────────────────────────────────────────────────────────────────────────
def export_result_3_from_workspace(db: Session) -> dict:
    from app.models.task3_workspace import Task3Workspace
    from app.models.task3_question_item import Task3QuestionItem
    from sqlalchemy import select, func

    stmt = select(Task3Workspace).order_by(Task3Workspace.id.desc()).limit(1)
    workspace = db.execute(stmt).scalar_one_or_none()

    if workspace is None:
        raise ValueError("工作台不存在，请先导入附件6")

    stmt = select(Task3QuestionItem).where(
        Task3QuestionItem.workspace_id == workspace.id
    ).order_by(Task3QuestionItem.question_code)
    questions = list(db.execute(stmt).scalars().all())

    if not questions:
        raise ValueError("没有可导出的题目")

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="结果汇总")
    ws.append(["编号", "问题", "SQL查询语句", "图形格式", "回答"])

    success_count = 0
    fail_count = 0

    for question in questions:
        question_id = question.question_code
        question_text = question.question_raw_json or ""
        rounds = _parse_question_rounds(question_text)
        try:
            answer_json = _ensure_non_empty_qa_pairs(question_text, question.answer_json or [])
            chart_type = question.chart_type or "无"

            ws.append([
                question_id,
                json.dumps(rounds, ensure_ascii=False),
                question.sql_text or "",
                chart_type,
                json.dumps(answer_json, ensure_ascii=False),
            ])

            success_count += 1

        except Exception as exc:
            logger.error("导出题目失败: question_id=%s, error=%s", question_id, str(exc))
            fail_count += 1

            ws.append([
                question_id,
                json.dumps(rounds, ensure_ascii=False),
                "",
                "无",
                json.dumps(
                    [{
                        "Q": rounds[0].get("Q", "") if rounds else question_text,
                        "A": {"content": f"处理失败: {str(exc)}"},
                    }],
                    ensure_ascii=False,
                ),
            ])

    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    result_dir = os.path.join(os.getcwd(), "result")
    os.makedirs(result_dir, exist_ok=True)

    result_path = os.path.join(result_dir, "result_3.xlsx")
    wb.save(result_path)

    # 更新工作台导出信息
    workspace.last_export_path = result_path
    workspace.last_exported_at = datetime.now()
    db.flush()

    logger.info(
        "result_3.xlsx 导出完成: path=%s, success=%d, failed=%d",
        result_path,
        success_count,
        fail_count,
    )

    return {
        "xlsx_path": result_path,
        "success_count": success_count,
        "fail_count": fail_count,
        "total": len(questions),
    }


def get_latest_export_info(db: Session) -> dict | None:
    from app.models.task3_workspace import Task3Workspace
    from sqlalchemy import select

    stmt = select(Task3Workspace).order_by(Task3Workspace.id.desc()).limit(1)
    workspace = db.execute(stmt).scalar_one_or_none()

    if workspace is None or not workspace.last_export_path:
        return None

    return {
        "xlsx_path": workspace.last_export_path,
        "exported_at": workspace.last_exported_at.isoformat() if workspace.last_exported_at else None,
        "total_questions": workspace.total_questions,
        "answered_count": workspace.answered_count,
    }
